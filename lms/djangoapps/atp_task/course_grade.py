# -*- coding: utf-8 -*-
import json
import logging
import os.path
import os
import time
import base64
from io import BytesIO

from opaque_keys.edx.keys import CourseKey

from util.json_request import JsonResponse
from django.http import Http404, HttpResponseServerError,HttpResponse
from openedx.core.djangoapps.site_configuration import helpers as configuration_helpers

from xmodule.modulestore.django import modulestore
from course_api.blocks.api import get_blocks
from lms.djangoapps.courseware.models import StudentModule
from course_api.blocks.views import BlocksInCourseView,BlocksView
from lms.djangoapps.course_blocks.api import get_course_blocks
from opaque_keys.edx.locations import SlashSeparatedCourseKey
from opaque_keys.edx.locator import CourseLocator
from courseware.courses import get_course_by_id
from courseware.views.views import _submission_history_context
from student.models import *
from django.contrib.auth.models import User
from importlib import reload

from student.models import UserPreprofile

from django.core.mail import EmailMessage

from lms.djangoapps.grades.course_grade_factory import CourseGradeFactory
from lms.djangoapps.persisted_grades.models import set_quiz_completion, get_persisted_course_grade

from django.conf import settings

from pprint import pformat
import datetime

import sys

log = logging.getLogger(__name__)

class course_grade():

    def __init__(self,course_id,course_key=None,request=None):

        self.course_id = course_id
        self.course_key = course_key
        self.request = request

    def get_titles(self):
        log.info("get_titles: Starting to get titles")
        if self.course_key is None:
            self.course_key = CourseLocator.from_string(self.course_id)


        #get all course structure
        course_usage_key = modulestore().make_course_usage_key(self.course_key)
        log.info(self)
        blocks = get_blocks(self.request,course_usage_key,depth='all',requested_fields=['display_name','children'])
        _root = blocks['root']
        blocks_overviews = []
        log.info("get_titles: Got course structure")
        #return unit title and component root
        try:
            children = blocks['blocks'][_root]['children']
            for z in children:
                child = blocks['blocks'][z]
                try:
                    sub_section = child['children']
                    for s in sub_section:
                        sub_ = blocks['blocks'][s]
                        vertical = sub_['children']
                        try:
                            for v in vertical:
                                unit = blocks['blocks'][v]
                                w = {}
                                w['id'] = unit['id']
                                w['display_name'] = unit['display_name']
                                try:
                                    w['children'] = unit['children']
                                except:
                                    pass
                                blocks_overviews.append(w)
                        except:
                            pass
                except:
                    pass
        except:
            pass
        log.info("get_titles: Got unit titles and component root")
        ## GET ALL SCORED XBLOCKS FOR WHICH THERE EXISTS AT LEAST ONE ENTRY IN STUDENTMODULE
        studentmodule = StudentModule.objects.raw("SELECT id,course_id,module_id FROM courseware_studentmodule WHERE course_id = %s AND max_grade IS NOT NULL AND grade <= max_grade GROUP BY module_id ORDER BY created", [self.course_id])

        title = []

        for n in studentmodule:
            try:
                usage_key = n.module_state_key
                _current = get_blocks(self.request,usage_key,depth='all',requested_fields=['display_name'])
                root = _current['root']

                unit_name = ''

                for over in blocks_overviews:
                    if str(root) in over.get('children'):
                        unit_name = over.get('display_name')


                q = {
                    "title":_current['blocks'][root]['display_name'],
                    "root":root,
                    'unit':unit_name
                }
                title.append(q)
            except:
                pass

        log.info("get_titles: All titles fetched")
        return title


    def _user(self,user_id):
        log.info("_user: starting for userid "+pformat(user_id))
        course_key = CourseLocator.from_string(self.course_id)
        course_block = StudentModule.objects.all().filter(student_id=user_id,course_id=course_key,max_grade__isnull=False)
        course_grade = []

        for n in course_block:
            try:
                q = {}
                usage_key = n.module_state_key
                block_name = get_blocks(self.request,usage_key,depth='all',requested_fields=['display_name'])
                root = block_name['root']
                display_name = block_name['blocks'][root]['display_name']
                q['earned'] = n.grade
                q['possible'] = n.max_grade
                q['display_name'] = display_name
                q['root'] = root
                course_grade.append(q)
            except:
                pass

        return course_grade


    def export(self,sended_email):
        log.warning("export: Start Task grade reports course_id : "+str(self.course_id) )
        course_key = CourseKey.from_string(self.course_id)
        course = get_course_by_id(course_key)
        course_enrollement = CourseEnrollment.objects.filter(course_id=course_key)

        #prepare xls
        header = [
            "id","email","first name","last name","level 1","level 2","level 3","level 4","status",'first access date','first success date'
        ]

        # Email content
        language_setup = {
            "en":{
                "subject": "Score report for {}",
                "text_content": "Please, find attached the score report for {}.\nRemember that if your training campaign is still in progress, this file is an intermediate statistical status."
            },
            "fr": {
                "subject": "Résultats des participants du module {}",
                "text_content": "Veuillez trouver en pièce attachée les résultats des participants pour le module {}.\nA noter que si votre campagne de formation est toujours en cours, ce fichier constitue un état statistique intermédiaire."
            }
        }

        title = self.get_titles()

        for n in title:
            header.append(n.get('unit')+' - '+n.get('title'))

        header.append('total grade (in %)')

        filename = '{}_grades_reports.xls'.format(self.course_id).replace('+','_')
        from openpyxl import Workbook
        ws = Workbook()
        virtual_workbook = BytesIO()
        sheet = ws.active
        
        for i, head in enumerate(header):
            i = i + 1
            sheet.cell(row=1,column=i).value = head
        j = 1

        for i in range(len(course_enrollement)):

            j = j + 1

            user=course_enrollement[i].user
            course_grade = CourseGradeFactory().read(user, course)

            user_id = user.id
            email = user.email
            first_name = user.first_name
            last_name = user.last_name
            log.info("export: getting grade for user: "+pformat(user.email))

            try:
                pre = UserPreprofile.objects.get(email=email)
                _lvl = [pre.level_1,pre.level_2,pre.level_3,pre.level_4]
            except:
                _lvl = ["","","",""]

            final_grade = course_grade.percent * 100

            progress_status = "started"
            passed = course_grade.passed
            percent = course_grade.percent

            course_id = self.course_id
            course_progression = 0

            from course_progress.helpers import get_overall_progress
            from openedx.features.course_experience.utils import get_course_outline_block_tree
            from completion.api.v1.views import SubsectionCompletionView

            total_blocks=0
            total_blockstma=0
            completed_blockstma=0
            completed_blocks=0
            completion_rate=0
            quiz_completion=0
            quiz_total_components=0
            quiz_completed_components=0
            quiz_completion_rate=0
            problem_ids = []
            course_sections = get_course_outline_block_tree(self.request,str(course_id)).get('children')
            if course_sections:
                for section in course_sections :
                  total_blockstma+=1
                  section_completion = SubsectionCompletionView().get(self.request,user,str(course_id),section.get('id')).data
                  if section.get('children') is None:
                    continue
                  for subsection in section.get('children') :
                    if subsection.get('children'):
                        for unit in subsection.get('children'):
                            total_blocks+=1
                            unit_completion = SubsectionCompletionView().get(self.request,user,str(course_id),unit.get('id')).data
                            if unit_completion.get('completion'):
                                completed_blocks+=1
                            if unit.get('num_graded_problems') > 1 and unit.get('graded') and unit.get('children'):
                                quiz_completion_rate = SubsectionCompletionView().get(self.request,user,str(course_id),unit.get('id')).data['completion']
                                for component in unit.get('children') :
                                    quiz_total_components+=1
                                    problem_ids.append(component.get('id'))
                                    if component.get('complete'):
                                        quiz_completed_components+=1
                        if completed_blocks == total_blocks:
                            completed_blockstma+=1

            if total_blocks != 0:
                course_progression = float(completed_blocks)/total_blocks

            if quiz_completion_rate == 1:
                set_quiz_completion(str(course_id),user.id)

            _end = 0
            try:
                _end = int(enrollment.course_overview.end.strftime("%s"))
            except:
                pass
            course_open = True
            if _end > 0 and _end < _now:
                course_open = False
            if course_open :
                if passed :
                    progress_status = "completed"
                else :
                    if course_progression == 0:
                        progress_status = "to do"
                    else:
                        progress_status = "in progress"
            else :
                progress_status = "completed"

            _user_blocks = self._user(user_id)

            sheet.cell(row=j, column=1).value = user_id
            sheet.cell(row=j, column=2).value = email
            sheet.cell(row=j, column=3).value = first_name
            sheet.cell(row=j, column=4).value = last_name
            sheet.cell(row=j, column=5).value = _lvl[0]
            sheet.cell(row=j, column=6).value = _lvl[1]
            sheet.cell(row=j, column=7).value = _lvl[2]
            sheet.cell(row=j, column=8).value = _lvl[3]
            sheet.cell(row=j, column=9).value = progress_status

            first_access_date = ''
            first_success_date = ''
            persisted_grade = get_persisted_course_grade(course_key, user_id)

            if persisted_grade:


                if persisted_grade.first_access:
                    first_access_date = persisted_grade.first_access
                else:
                    # There is no first access date, still if course progress is >0 then we will "force" the first_access date if possible
                    # we do that because we have long history where this data was not computed
                    if course_progression > 0:
                        # Of course we bother only if course_progression is > 0
                        # Now let's find if there is any question for which the student answered
                        last_submission_times = []
                        for problem_id in problem_ids:
                            history_entries = _submission_history_context(user, course_id, user.username, problem_id).get('history_entries')
                            if history_entries:
                                for entry in history_entries:
                                    last_submission_times.append(entry.updated)
                                # We take the earliest submission made and remove 1 hour
                        first_access_date = min(last_submission_times) - datetime.timedelta(hours=1, minutes=0)
                if isinstance(first_access_date, datetime.date):
                   # We try to save the date if possible as some other operations on database may be currently done
                    try:
                        persisted_grade.first_access = first_access_date
                        persisted_grade.save()
                    except:
                        pass
                    sheet.cell(row=j, column=10).value = first_access_date.strftime("%Y-%m-%d %H:%M")



                if persisted_grade.first_success:
                    first_success_date = persisted_grade.first_success
                elif passed:
                    # There is no first success date but we will find the latest date at which learner answered
                    last_submission_times = []
                    for problem_id in problem_ids:
                        history_entries = _submission_history_context(user, course_id, user.username, problem_id).get('history_entries')
                        if history_entries:
                            for entry in history_entries:
                                last_submission_times.append(entry.updated)
                    first_success_date = max(last_submission_times)
                if isinstance(first_success_date, datetime.date):
                    # We try to save the date if possible as some other operations on database may be currently done
                    try:
                        persisted_grade.first_success = first_success_date
                        persisted_grade.save()
                    except:
                        pass
                    sheet.cell(row=j, column=11).value = first_success_date.strftime("%Y-%m-%d %H:%M")

            k = 12



            for val in title:
                _grade = 0
                for block in _user_blocks:

                    if block.get('root') == val.get('root'):
                        _grade = block.get('earned')


                sheet.cell(row=j, column=k).value = _grade

                k = k + 1

            sheet.cell(row=j, column=k).value =final_grade

        ws.save(virtual_workbook)
        _files_values = virtual_workbook.getvalue()

        log.warning("End Task grade reports course_id : "+str(self.course_id))

        #sending mail
        log.warning("send grade reports course_id : "+str(filename))
        log.warning("email5 : "+str(sended_email))

        # Email content according to course language (look for #Email content to add new languages)
        subject = ''
        text_content = ''
        if course.language in language_setup:
            subject = language_setup[course.language]['subject'].format(course.display_name_with_default_escaped)
            text_content = language_setup[course.language]['text_content'].format(course.display_name_with_default_escaped)
        else:
            subject = language_setup['en']['subject'].format(course.display_name_with_default_escaped)
            text_content = language_setup['en']['text_content'].format(course.display_name_with_default_escaped)

        from_email=configuration_helpers.get_value('email_from_address', settings.DEFAULT_FROM_EMAIL)
        to = sended_email
        mimetype='application/vnd.ms-excel'
        fail_silently=False
        _data = _files_values
        #_encoded = base64.b64encode(wb)
        log.warning("end send grade : right before email is sent out")
        log.warning(settings.DEFAULT_FROM_EMAIL)
        _email = EmailMessage(subject, text_content, from_email, [to])
        _email.attach(filename, _data, mimetype=mimetype)
        _email.send(fail_silently=fail_silently)

        log.warning("end send grade reports course_id : "+str(filename))

        context = {
            "filename" : filename
        }

        return context



    def get_xls(self):
        log.info("get_xls: starting")
        context = {
            "status": False,
            'filepath': None
        }

        folder_path = '/edx/var/edxapp/grades/'
        filename = '{}_grades_reports.xls'.format(self.course_id).replace('+','_')
        filepath = folder_path+filename
        if os.path.isfile(filepath):
            context['status'] = True,
            context['filepath'] = filepath
            context['time'] = time.ctime(os.path.getmtime(filepath))
            log.info("get_xls: done")
        log.info("get_xls")
        return context

    def download_xls(self,filename):
        log.info("download_xls: starting")
        full_path = '/edx/var/edxapp/grades/'+filename
        _file = open(full_path,'r')
        _content = _file.read()
        response = HttpResponse(_content, content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = "attachment; filename="+filename
        os.remove(full_path)
        log.info("download_xls: done")
        return response
