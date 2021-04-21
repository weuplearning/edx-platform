# -*- coding: utf-8 -*-
import sys
import importlib
importlib.reload(sys)

import os
import json
# from xlwt import *
import time
import logging
import re

from django.utils.translation import ugettext as _

from django.conf import settings

from django.http import Http404, HttpResponseServerError, HttpResponse
from util.json_request import JsonResponse
from student.models import User,CourseEnrollment,UserProfile
from openedx.core.djangoapps.site_configuration import helpers as configuration_helpers
# from lms.djangoapps.tma_grade_tracking.models import dashboardStats
from lms.djangoapps.wul_apps.ensure_form.utils import ensure_form_factory
from .libs import return_select_value
# from lms.djangoapps.grades.new.course_grade import CourseGradeFactory
# from opaque_keys.edx.locations import SlashSeparatedCourseKey
from opaque_keys.edx.keys import CourseKey
from courseware.courses import get_course_by_id
# from openedx.core.djangoapps.course_groups.models import CohortMembership, CourseUserGroup
from openedx.core.djangoapps.course_groups.cohorts import get_cohort, is_course_cohorted
from lms.djangoapps.wul_apps.models import WulCourseEnrollment
import time
from collections import OrderedDict
# from lms.djangoapps.grades.context import grading_context_for_course

from io import BytesIO

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from django.core.mail import EmailMessage
# from microsite_configuration.models import Microsite
from django.conf import settings
from lms.djangoapps.wul_apps.best_grade.helpers import check_best_grade
from courseware.user_state_client import DjangoXBlockUserStateClient
from openpyxl import Workbook as openpyxlWorkbook

log = logging.getLogger(__name__)

#OLD VERSION GRADES REPORT
class grade_reports():
    def __init__(self,request,course_id=None,microsite=None,filename=None,filepath=None,subscribe_report=False):
        self.request = request
        self.course_id = course_id
        self.microsite = microsite
        self.filename = filename
        self.filepath = filepath
        self.subscribe_report = subscribe_report

    # def prepare_workbook(self):
    #     timesfr = time.strftime("%d_%m_%Y_%H_%M_%S")
    #     timestr = str(timesfr)
    #     wb = Workbook(encoding='utf-8')
    #     self.filename = '{}_{}_grades.xls'.format(self.microsite,timestr)
    #     self.filepath = '/edx/var/edxapp/grades/{}'.format(self.filename)
    #     sheet_count = wb.add_sheet('reports')

    #     return wb

    # def generate_xls(self):
    #     log.warning(_("tma_grade_reports : start generate_xls"))
    #     log.warning(_("tma_grade_reports : course_id : "+self.course_id))

    #     #get request body

    #     body = json.loads(self.request.body).get('fields')

    #     log.warning(_("tma_grade_reports : microsite : "+self.microsite))
    #     log.warning(_("tma_grade_reports : body : "+str(body)))
    #     #get mongodb course's grades
    #     mongo_persist = dashboardStats()
    #     find_mongo_persist_course = mongo_persist.get_course(self.course_id)

    #     #prepare workbook
    #     _wb = self.prepare_workbook()
    #     sheet_count = _wb.get_sheet('reports')

    #     #form factory connection
    #     form_factory = ensure_form_factory()
    #     form_factory_db = 'ensure_form'
    #     form_factory_collection = 'certificate_form'
    #     form_factory.connect(db=form_factory_db,collection=form_factory_collection)

    #     #get register & certificate fields info
    #     register_form_ = configuration_helpers.get_value('FORM_EXTRA')
    #     certificates_form_ = configuration_helpers.get_value('CERTIFICATE_FORM_EXTRA')

    #     #prepare workbook header
    #     k=0
    #     for u in body:
    #         if not ('final-grades' and 'summary') in u:
    #             sheet_count.write(0, k, u)
    #             k = k + 1
    #         elif "summary" in u:
    #             grade_breakdown = find_mongo_persist_course[0].get('users_info').get('summary').get('grade_breakdown')
    #             for key,value in grade_breakdown.items():
    #                 sheet_count.write(0, k, value.get('category'))
    #                 k = k + 1
    #         elif "final-grades" in u:
    #             sheet_count.write(0, k, 'final grade')
    #             k = k + 1
    #             sheet_count.write(0, k, 'Eligible attestation')
    #             k = k + 1

    #     #sheets row
    #     j = 1

    #     #write differents workbook_rows
    #     for _row in find_mongo_persist_course:
    #         # get row user_id
    #         user_id = _row['user_id']

    #         #ensure user exist if False pass to next row
    #         ensure_user_exist = True
    #         try:
    #             #get current row users mysql auth_user & aut_userprofile info
    #             _user = User.objects.raw("SELECT a.id,a.email,a.username,a.date_joined,b.custom_field FROM auth_user a,auth_userprofile b WHERE a.id = b.user_id and a.id = %s", [user_id])

    #             #prepare dict of auth_user info
    #             user = {}

    #             #prepare dict of aut_userprofile info
    #             user_profile = {}

    #             #prepare dict of certificate form info
    #             user_certif_profile = {}

    #             #ensure only first occurance of _user is use
    #             det = 0

    #             #get current user certificates forms values
    #             if certificates_form_ is not None:
    #                 form_factory.microsite = self.microsite
    #                 form_factory.user_id = user_id
    #                 try:
    #                     user_certif_profile = form_factory.getForm(user_id=True,microsite=True).get('form')
	# 		log.ingo("user_certif_profile")
	# 		log.warning(user_certif_profile)
    #                 except:
    #                     pass

    #             #hydrate user & user_profile dicts
    #             for extract in _user:
    #                 if det < 1:
    #                     #hydrate user dict
    #                     user['id'] = extract.id
    #                     user['email'] = extract.email
    #                     user['username'] = extract.username
    #                     user['date_joined'] = extract.date_joined

    #                     #row from auth_userprofile mysql table
    #                     _user_profile = extract.custom_field

    #                     #bloc for with det == 1
    #                     det = 1

    #                     #hydrate user_profile dict
    #                     try:
    #                         user_profile = json.loads(_user_profile)
    #                     except:
    #                         pass

    #         except:
    #             ensure_user_exist = False

    #         #write user xls line if exist
    #         if ensure_user_exist:
    #             k=0
    #             for n in body:

    #                 #insert user mysql value to xls
    #                 if n in user.keys():
    #                     if n == 'date_joined':
    #                         sheet_count.write(j, k, str(user.get(n).strftime('%d-%m-%Y')))
    #                     else:
    #                         sheet_count.write(j, k, user.get(n))
    #                     k = k + 1

    #                 #insert register_form mysql value to xls
    #                 elif n in user_profile.keys():
    #                     _insert_value = return_select_value(n,user_profile.get(n),register_form_)
    #                     sheet_count.write(j, k, _insert_value)
    #                     k = k + 1

    #                 #insert certificate_form mongodb value to xls
    #                 elif n in user_certif_profile.keys():
    #                     _insert_value = return_select_value(n,user_certif_profile.get(n),certificates_form_)
    #                     sheet_count.write(j, k, _insert_value)
    #                     k = k + 1

    #                 #insert summary grades mongodb value to xls
    #                 elif "summary" in n:
    #                     grade_breakdown = _row.get('users_info').get('summary').get('grade_breakdown')
    #                     for key,value in grade_breakdown.items():
    #                         #insert grade value to xls
    #                         details = value['detail']
    #                         details = details.replace(value['category'],"").replace(" = ","").replace("of a possible ","").replace("%","")
    #                         split = details.split(" ")
    #                         avg = str(int(float(split[0])/float(split[1]) * 100))+"%"
    #                         sheet_count.write(j, k, avg)
    #                         k = k + 1

    #                 #insert final grades mongodb value to xls
    #                 elif "final-grades" in n:
    #                     sheet_count.write(j, k, str(_row.get('users_info').get('percent') * 100)+"%")
    #                     k = k + 1
    #                     sheet_count.write(j, k, str(_row.get('users_info').get('passed')))
    #                     k = k + 1
    #                 else:
    #                     sheet_count.write(j, k, '')
    #                     k = k + 1
    #             j = j + 1
    #         else:
    #             pass
    #     log.warning(_("tma_grade_reports : save file generate_xls"))
    #     _wb.save(self.filepath)
    #     _file = open(self.filepath,'r')
    #     _content = _file.read()
    #     _file.close()

    #     response = {
    #         'path':self.filename
    #     }
    #     return JsonResponse(response)




    # def tma_get_scorable_blocks_titles(self, course_key):
    #     """
    #     Returns an dict that maps a scorable block's location id to its title.
    #     """
    #     scorable_block_titles = OrderedDict()
    #     grading_context = grading_context_for_course(course_key)

    #     for assignment_type_name, subsection_infos in grading_context['all_graded_subsections_by_type'].iteritems():
    #         for subsection_index, subsection_info in enumerate(subsection_infos, start=1):
    #             for scorable_block in subsection_info['scored_descendants']:
    #                 header_name = (
    #                     u"{assignment_type} {subsection_index}: "
    #                     u"{subsection_name} - {scorable_block_name}"
    #                 ).format(
    #                     scorable_block_name=scorable_block.display_name,
    #                     assignment_type=assignment_type_name,
    #                     subsection_index=subsection_index,
    #                     subsection_name=subsection_info['subsection_block'].display_name,
    #                 )
    #                 scorable_block_titles[scorable_block.location] = header_name
    #     return scorable_block_titles

    # def get_time_tracking(self,enrollment):
    #     tma_enrollment,is_exist=TmaCourseEnrollment.objects.get_or_create(course_enrollment_edx=enrollment)
    #     global_time_tracking = self.convert_seconds_in_hours(tma_enrollment.global_time_tracking)
    #     return global_time_tracking
        

    # def get_detailed_time_tracking(self,enrollment):
    #     detailed_time_tracking = {}
    #     tma_enrollment,is_exist=TmaCourseEnrollment.objects.get_or_create(course_enrollment_edx=enrollment)
    #     try:
    #         detailed_time_tracking = json.loads(tma_enrollment.detailed_time_tracking)
    #     except:
    #         pass
    #     return detailed_time_tracking

    # def convert_seconds_in_hours(self, seconds):
    #     hours = seconds // 3600
    #     seconds %= 3600
    #     minutes = seconds // 60
    #     time_spent = str(hours)+"h"+str(minutes)+"min"
    #     return time_spent

        
        
    #ACTUAL VERSION GRADES REPORT
    def task_generate_xls(self):
        #Get report infos
        self.microsite = self.request.get('microsite')
        report_fields = self.request.get('form')
        register_fields = self.request.get('register_form')
        certificate_fields = self.request.get('certificate_form')
        select_custom_field_key=self.request.get('select_custom_field_key')
        select_custom_field_values=self.request.get('select_custom_field_values')
        split_by=self.request.get('split_by')
        do_not_send_email=self.request.get('do_not_send_email',False)
        scope = self.request.get('scope')
        users_admin_list = self.request.get('users_admin')
        include_days_left_in_report = self.request.get('include_days_left_in_report')
        certificate_advanced_config = self.request.get('certificate_advanced_config')

        log.info('Start task generate grade report for course {}'.format(self.course_id))

        course_key=CourseKey.from_string(self.course_id)
        course=get_course_by_id(course_key)
        microsite_information = Microsite.objects.get(key=self.microsite)

        form_factory = ensure_form_factory()
        form_factory.connect(db='ensure_form',collection='certificate_form')

        #Dict of labels
        form_labels={
            "last_connexion":_("Last login"),
            "inscription_date":_("Register date"),
            "user_id":_("User id"),
            "email":_("Email"),
            "grade_final":_("Final Grade"),
            "cohorte_names":_("Cohorte name"),
            "time_tracking":_("Time spent"),
            "certified":_("Attestation"),
            "username":_("Username"),
            "best_grade":_("Best Grade"),
            "best_grade_date":_("Best Grade Date"),
            "detailed_time_tracking": _("Time spent"),
        }

        if include_days_left_in_report :    
            form_labels["days_left"]="Nombre de jours restants"
    

        for field in register_fields :
            form_labels[field.get('name')]=field.get('label')
        for field in certificate_fields :
            form_labels[field.get('name')]=field.get('label')

        calculate_average_for_attestation = False
        if form_labels["cas_pratique_grade"]:
            if not certificate_advanced_config:
                form_labels.pop("cas_pratique_grade")
            elif certificate_advanced_config and len(certificate_advanced_config) == 0:
                form_labels.pop("cas_pratique_grade")
            elif certificate_advanced_config and len(certificate_advanced_config) > 0 and not str(course_key) in certificate_advanced_config:
                form_labels.pop("cas_pratique_grade")
            else:
                calculate_average_for_attestation = True

        #Identify multiple cells fields
        multiple_cell_fields=["exercises_grade","grade_detailed","exercises_answers"]

        #Is report cohort specific?
        course_cohorted=is_course_cohorted(course_key)
        cohortes_targeted = []
        if course_cohorted :
            cohortes_targeted=[field.replace('cohort_selection_','') for field in report_fields if field.find('cohort_selection_')>-1]
            if cohortes_targeted and not 'cohorte_names' in report_fields:
                report_fields.append('cohorte_names')
        else :
            if 'cohorte_names' in report_fields:
                report_fields.remove('cohorte_names')

        #Get Graded block for exercises_grade details
        scorable_blocks_titles = self.tma_get_scorable_blocks_titles(course_key)

        #Create Workbook
        wb = openpyxlWorkbook()
        filename = '/home/edxtma/csv/{}_{}.xls'.format(time.strftime("%Y_%m_%d"),course.display_name_with_default.replace("\"","").replace("\'",""))
        sheet =  wb.active
        sheet.title = "Grade Report"

        #Get course chapters for detailed time tracking report generation
        course_chapters = []
        for chapter in course.get_children():
            course_chapters.append({
                "id": chapter.url_name,
                "name": chapter.display_name_with_default_escaped,
            })

        #Write information
        line=1
        course_enrollments=CourseEnrollment.objects.filter(course_id=course_key, is_active=1)

        for enrollment in course_enrollments:
            try:
                include_admin = scope["admin"]
                include_intern = scope["intern"]
            except:
                pass

            user= enrollment.user
            log.info("-------------------------- treating user {} for grade report -------------------------".format(user.email))


            try:
                #case 1 : include admin and intern : nothing change 

                #case 2 : include intern but not admin :
                if not include_admin and user.email in users_admin_list:
                    continue
                elif not include_admin and hasattr(users_admin_list, '__iter__'):
                    if any(re.match(pattern + "$", user.email) for pattern in users_admin_list):
                        continue

                #case 3 : include admin but not intern :
                if not include_intern and hasattr(users_admin_list, '__iter__'):
                    if any(re.match(pattern + "$", user.email) for pattern in users_admin_list):
                        pass
                    elif not include_intern and user.email not in users_admin_list:
                        continue
            except:
                pass

            #do not include in reports if not active
            if not enrollment.is_active:
                continue

            #Cohort Specific report - remove students not targetted
            if course_cohorted :
                user_cohorte=get_cohort(user, course_key).name
                if cohortes_targeted and not user_cohorte in cohortes_targeted :
                    continue

            #Custom field specific report - remove students not targetted
            #Get user custom field
            try:
                custom_field = json.loads(UserProfile.objects.get(user=user).custom_field)
            except:
                custom_field = {}
            if select_custom_field_key :
                if select_custom_field_values and not custom_field.get(select_custom_field_key) in select_custom_field_values :
                    continue
                else :
                    if not select_custom_field_key in report_fields:
                        report_fields.append(select_custom_field_key)

            #Gather user information
            user_grade = []
            grade_summary={}
            if not self.subscribe_report:
                user_grade = check_best_grade(user, course, force_best_grade=("best_grade" in report_fields))
                for section_grade in user_grade.grade_value['section_breakdown']:
                    grade_summary[section_grade['category']]=section_grade['percent']

            tma_enrollment=TmaCourseEnrollment.get_enrollment(course_id=self.course_id, user=user)
            user_certificate_info = {}
            try:
                form_factory.microsite = self.microsite
                form_factory.user_id = user.id
                user_certificate_info = form_factory.getForm(user_id=True,microsite=True).get('form')
            except:
                pass

            user_state_client = DjangoXBlockUserStateClient()

            cell=1

            #get the number of days left
            milliseconds_in_a_day = 86400000
            milliseconds_in_a_month = 2629743000
            days_left_in_days = 0
            user_date_from_custom_fields = 0
            try : 
                user_date_from_custom_fields =  custom_field.get(str(course_key))
                user_date_in_milliseconds = (int(user_date_from_custom_fields) + milliseconds_in_a_month) / 1000
                date_today_in_milliseconds = time.time() 

                if  user_date_in_milliseconds > date_today_in_milliseconds :
                    days_left_in_milliseconds = (user_date_in_milliseconds - date_today_in_milliseconds)

                    days_left_in_days = (days_left_in_milliseconds / milliseconds_in_a_day) * 1000
            except:
                pass

            #get cas_pratique_grade
            form_factory.get_user_form_extra(user)
            form_factory.get_user_certificate_form_extra(user)
            certificate_form_extra = form_factory.user_certificate_form_extra
            user_cas_pratique_grade = 0
            try:
                user_cas_pratique_grade = certificate_form_extra["cas_pratique_grade"]
            except:
                pass


            # average mark between cas_pratique_grade and best_grade
            if calculate_average_for_attestation:
                cas_pratique_grade_coefficient = certificate_advanced_config[str(course_key)]['weight']
                best_grade_coefficient = 100 - cas_pratique_grade_coefficient
                user_best_grade = int(round(user_grade.percent_tma*100))
                average_mark = ((cas_pratique_grade_coefficient* int(user_cas_pratique_grade) + best_grade_coefficient * user_best_grade) / 100)




            for field in report_fields :
                if field in multiple_cell_fields:
                    if field=="grade_detailed":
                        for section in sorted(grade_summary):
                            section_grade = str(int(round(grade_summary[section] * 100)))+'%'
                            sheet.cell(row=line+1, column=cell, value=section_grade)
                            if line ==1 :
                                sheet.cell(row=1, column=cell, value="Travail - "+section)
                            cell+=1
                    elif field=="exercises_grade":
                        for block_location,block_title  in scorable_blocks_titles.items():
                            try:
                                if user_grade.locations_to_scores.get(block_location):
                                    block_detail = user_grade.locations_to_scores.get(block_location)
                                    if block_detail.attempted:
                                        value=round(float(block_detail.earned)/block_detail.possible, 2)
                                    else:
                                        value=_('n.a.')
                                else :
                                    value=('not graded for student')
                            except:
                                value=_('inv.')
                            sheet.cell(row=line+1, column=cell, value=value)
                            if line==1 :
                                sheet.cell(row=1, column=cell, value="Grade - "+block_title)
                            cell+=1
                            if "exercises_answers" in report_fields:
                                #Answer
                                try:
                                    if user_grade.locations_to_scores.get(block_location):
                                        history_entries = list(user_state_client.get_history(user.username, block_location))
                                        value = history_entries[0].state.get('student_answers').values()[0]
                                        if isinstance(value, basestring) and "choice_" in value:
                                            value=self.addOneToChoice(value)
                                        elif isinstance(value, basestring) and not "choice_" in value:
                                            value = "diff type"
                                        elif isinstance(value, list):
                                            value=[self.addOneToChoice(choice) for choice in value]
                                            value=", ".join(value)
                                    else:
                                        value=('not graded for student')
                                except:
                                    value=_('inv.')
                                sheet.cell(row=line+1, column=cell, value=value)
                                if line==1 :
                                    sheet.cell(row=1, column=cell, value="Answers - "+scorable_blocks_titles[block_location])
                                cell+=1
                                #Submission Stamp
                                try:
                                    if user_grade.locations_to_scores.get(block_location) and history_entries:
                                        value=history_entries[0].state.get('last_submission_time')
                                    else:
                                        value=('no time stamp')
                                except:
                                    value=_('inv.')
                                sheet.cell(row=line+1, column=cell, value=value)
                                if line==1 :
                                    sheet.cell(row=1, column=cell, value="Last submission - "+scorable_blocks_titles[block_location])
                                cell+=1
                                history_entries=None

                    elif field=="exercises_answers" and not "exercises_grade" in report_fields:
                        for block_location,block_title  in scorable_blocks_titles.items():
                            #Answer
                            try:
                                if user_grade.locations_to_scores.get(block_location):
                                    history_entries = list(user_state_client.get_history(user.username, block_location))
                                    value = history_entries[0].state.get('student_answers').values()[0]
                                    if isinstance(value, basestring) and "choice_" in value:
                                        value=self.addOneToChoice(value)
                                    elif isinstance(value, basestring) and not "choice_" in value:
                                        value = "diff type"
                                    elif isinstance(value, list):
                                        value=[self.addOneToChoice(choice) for choice in value]
                                        value=", ".join(value)
                                else:
                                    value=('not graded for student')
                            except:
                                value=_('inv.')
                            sheet.cell(row=line+1, column=cell, value=value)
                            if line==1 :
                                sheet.cell(row=1, column=cell, value="Answers - "+scorable_blocks_titles[block_location])
                            cell+=1
                            #Submission Stamp
                            try:
                                if user_grade.locations_to_scores.get(block_location) and history_entries:
                                    value=history_entries[0].state.get('last_submission_time')
                                else:
                                    value=('no time stamp')
                            except:
                                value=_('inv.')
                            sheet.cell(row=line+1, column=cell, value=value)
                            if line==1 :
                                sheet.cell(row=1, column=cell, value="Last submission - "+scorable_blocks_titles[block_location])
                            cell+=1
                            history_entries=None
                else :
                    value = ''
                    values = {}
                    if field=="user_id":
                        value=user.id
                    elif field=="email":
                        value=user.email
                    elif field=="first_name":
                        try :
                            if user.first_name:
                                value=user.first_name
                            elif custom_field :
                                value=custom_field.get('first_name', 'unkowna')
                            else :
                                value='unknown'
                        except :
                            value='unknown'
                    elif field=="last_name":
                        try :
                            if user.last_name:
                                value=user.last_name
                            elif custom_field:
                                value=custom_field.get('last_name', 'unkowna')
                        except :
                            value='unknown'
                    elif field=="last_connexion":
                        try :
                            value=user.last_login.strftime('%d-%m-%y')
                        except:
                            value=''
                    elif field=="inscription_date":
                        try :
                            value=user.date_joined.strftime('%d-%m-%y')
                        except:
                            value=''
                    elif field=="cohorte_names":
                        try:
                            value=user_cohorte
                        except:
                            value=''
                    elif field=="time_tracking":
                        value=self.get_time_tracking(enrollment)
                    elif field=="best_grade":
                        try:
                            value = str(int(round(user_grade.percent_tma*100)))+"%"
                        except:
                            value='n.a'
                    elif field=="best_grade_date":
                        try:
                            value=tma_enrollment.best_grade_date.strftime('%d-%m-%y')
                        except:
                            value='n.a'
                    elif field=="certified":
                        if calculate_average_for_attestation:
                            if average_mark > 60:
                                value = _("Yes")
                            else :
                                value = _("No")
                        else:
                            if user_grade.passed :
                                value = _("Yes")
                            else :
                                value = _("No")
                    elif field=="grade_final":
                        value = str(int(round(user_grade.percent * 100)))+'%'
                    elif field=="username":
                        value=user.username
                    elif field in user_certificate_info.keys():
                        value=user_certificate_info.get(field)
                    elif field=="detailed_time_tracking":                    
                        values = self.get_detailed_time_tracking(enrollment)
                    elif field=="days_left":              
                        value = round(days_left_in_days)
                    elif field=="cas_pratique_grade":
                        value = str(user_cas_pratique_grade) + "%"

                    else :
                        value=custom_field.get(field,'')

                    #Write header and write value
                    if field in form_labels.keys():
                        if field != "detailed_time_tracking":
                            sheet.cell(row=line+1, column=cell, value=value)
                            if line==1:
                                sheet.cell(row=1, column=cell, value=form_labels.get(field))
                            cell+=1
                        else:
                            if line==1:
                                sheet.cell(row=1, column=cell, value=_("Time spent"))
                            sheet.cell(row=line+1, column=cell, value=self.get_time_tracking(enrollment))
                            cell+=1

                            for index, chapter in enumerate(course_chapters):
                                value = 0
                                if line==1:
                                    sheet.cell(row=1, column=cell+index, value=chapter["name"])
                                if chapter["id"] in values.keys():
                                    value = values[chapter["id"]]
                                sheet.cell(row=line+1, column=cell+index, value=self.convert_seconds_in_hours(value))    
                            cell+=len(course_chapters)

            line+=1

        #Save the file
        output = BytesIO()
        wb.save(output)
        _files_values = output.getvalue()

        #If no email is to be sent, then just send the response to the caller as a BytesIO object
        if do_not_send_email:
            response = {
                'path':self.filename,
                'send_to':receivers,
                'xls_file':_files_values
            }
            return response

        #Send the email to receivers
        receivers = self.request.get('send_to')

        if cohortes_targeted and len(cohortes_targeted)>1:
            html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en PJ le rapport de donnees du MOOC {} pour les cohortes {}<br/><br/>Si vous disposez d'accès suffisants vous pouvez accéder au dashboard du cours: https://{}/tma/{}/dashboard <br><br> et au studio du cours : https://{}/course/{}    <br/><br/>Bonne reception<br>The MOOC Agency<br></p></body></html>".format(course.display_name, ' , '.join(cohortes_targeted), microsite_information.values['SITE_NAME'], course.id, settings.CMS_BASE, course.id)
        elif cohortes_targeted and len(cohortes_targeted)==1:
            html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en PJ le rapport de donnees du MOOC {} pour la cohorte {}<br/><br/>Si vous disposez d'accès suffisants vous pouvez accéder au dashboard du cours: https://{}/tma/{}/dashboard <br><br> et au studio du cours : https://{}/course/{}    <br/><br/>Bonne reception<br>The MOOC Agency<br></p></body></html>".format(course.display_name, ' '.join(cohortes_targeted), microsite_information.values['SITE_NAME'], course.id, settings.CMS_BASE, course.id)
        else :
            html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en PJ le rapport de donnees du MOOC {}<br/><br/>Si vous disposez d'accès suffisants vous pouvez accéder au dashboard du cours: https://{}/tma/{}/dashboard <br><br> et au studio du cours : https://{}/course/{}    <br/><br/>Bonne reception<br>The MOOC Agency<br></p></body></html>".format(course.display_name, microsite_information.values['SITE_NAME'], course.id, settings.CMS_BASE, course.id)
        part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')

        for receiver in receivers :
            fromaddr = "ne-pas-repondre@themoocagency.com"
            toaddr = str(receiver)
            msg = MIMEMultipart()
            msg['From'] = fromaddr
            msg['To'] = toaddr

            subject = u"{} - Rapport de donnees".format(course.display_name)
            if cohortes_targeted:
                subject += ' - Filtre Cohortes :'+' '.join(cohortes_targeted)
            if select_custom_field_key and select_custom_field_values:
                subject += (' - Filtre : '+' '.join(select_custom_field_values))

            msg['Subject'] = subject

            attachment = _files_values
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
            msg.attach(part)
            server = smtplib.SMTP('mail3.themoocagency.com', 25)
            server.starttls()
            server.login('contact', 'waSwv6Eqer89')
            msg.attach(part2)
            text = msg.as_string()
            server.sendmail(fromaddr, toaddr, text)
            server.quit()
            log.warning("file sent to {}".format(receiver))

        response = {
            'path':self.filename,
            'send_to':receivers
        }

        return response

    # def addOneToChoice(self, choice):
    #     choice=choice.split("_")
    #     choice[1]=str(int(choice[1])+1)
    #     return "_".join(choice)

    # def download_xls(self):
    #     self.filepath = '/edx/var/edxapp/grades/{}'.format(self.filename)
    #     _file = open(self.filepath,'r')
    #     _content = _file.read()
    #     response = HttpResponse(_content, content_type="application/vnd.ms-excel")
    #     response['Content-Disposition'] = "attachment; filename="+self.filename
    #     return response
