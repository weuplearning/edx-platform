# -*- coding: utf-8 -*-
'''
Converter xls to targz courses feature

/edx/app/edxapp/edx-platform/lms/djangoapps/wul_apps/converter_xlsx_to_targz/bvt/
'''

from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
import json
import os
import xlrd
import tarfile
import xlwt

from lxml import etree
from openpyxl import load_workbook
import shutil
from datetime import date

import logging
log = logging.getLogger()


PROBLEMSHEET = "problem"
PROBLEMINDEX = 0
PROBLEMSECTION = 1
PROBLEMSUBSECTION = 2
PROBLEMUNIT = 3
PROBLEMDIR = 4
PROBLEMNAME = 5
PROBLEMSHEETNAME = 6
PROBLEMDISPLAYNAME = 7
PROBLEMTYPE = 8


class Problem_checkbox:

    def __init__(self, info):
        wb_prob = xlrd.open_workbook(
            '/edx/var/edxapp/media/microsites/bvt/edx_converter/problem_source/questions.xls')
        self.sheetstruc = wb_prob.sheet_by_name(info['sheet'])
        self.n_checkbox = (self.sheetstruc.ncols-4)//3
        self.prob_disp = 0
        self.prob_weight = 1
        self.prob_attempt = 2
        self.prob_hint = 3
        self.question_col = 4
        self.checkbox_col = 5
        self.ans_col = 6
        self.img = 7
        print('number of question is '+str(self.n_checkbox))
        if self.n_checkbox is float:
            print('number of column does not match with number of droplist')
            exit()

    def display_name(self):
        return(self.sheetstruc.cell_value(1, self.prob_disp))

    def hint(self):
        return(self.sheetstruc.cell_value(1, self.prob_hint))

    def weight(self):
        weigth_per_question = self.sheetstruc.cell_value(1, self.prob_weight)
        if weigth_per_question == '':
            return('')
        else:
            total_weight = float(float(weigth_per_question)*self.n_checkbox)
            return(str(total_weight))

    def attempt(self):
        if self.sheetstruc.cell_value(1, self.prob_attempt) == '':
            return('')
        else:
            attempt = int(self.sheetstruc.cell_value(1, self.prob_attempt))
            return(str(attempt))

    def checkbox(self, element_obj):

        for checkboxs_idx in range(0, self.n_checkbox):
            for row_ in range(1, self.sheetstruc.nrows):
                tmp = self.sheetstruc.cell_value(row_, self.question_col)
                if tmp != '':
                    question_page = etree.SubElement(element_obj, 'p')
                    question_page.text = tmp

            if self.sheetstruc.cell_value(1,self.img) == "" or self.sheetstruc.cell_value(1,self.img) == 0:
                log.info("image : no")
            else:
                log.info("image: yes")
                img_tag=etree.SubElement(element_obj,'img', src=self.sheetstruc.cell_value(1,self.img))
                img_tag.set('style', "display: block; margin: auto; max-height: 500px;")

            choice_response_page = etree.SubElement(element_obj,'choiceresponse')
            checkbox_group_page= etree.SubElement(choice_response_page,'checkboxgroup')
            for row_ in range(1,self.sheetstruc.nrows):
                answer_text = self.sheetstruc.cell_value(row_,self.ans_col)
                checkbox_text = self.sheetstruc.cell_value(row_,self.checkbox_col)

                if answer_text == '':
                    continue

                if answer_text.lower() == 't'.lower():
                    checkbox_obj = etree.SubElement(checkbox_group_page,'choice',correct='True')
                    checkbox_obj.text = checkbox_text
                else:
                    checkbox_obj = etree.SubElement(checkbox_group_page,'choice',correct='False')
                    checkbox_obj.text = checkbox_text

            self.question_col = self.question_col + 3
            self.checkbox_col = self.checkbox_col+3
            self.ans_col = self.ans_col+3


        if self.hint() != '':
            demand_hint = etree.SubElement(element_obj,'demandhint') 
            hint = etree.SubElement(demand_hint,'hint')
            hint.text = self.hint() 

            
        return(element_obj)


    def create_file(self,filename,course_path):
        new_problem_file = os.path.join(course_path,'problem',filename)

        page = etree.Element('problem', display_name=self.display_name()) 
        if self.weight() != '':
            page.set('weight',self.weight())

        if self.attempt() != '':
            page.set('max_attempts',self.attempt())

        full_xml_obj = self.checkbox(page)
        doc = etree.ElementTree(page)
        doc.write(new_problem_file, pretty_print=True, xml_declaration=False, encoding='utf-8')


def problem_excel2list(row,sheetproblem):

    problem_info = []
    problem_idx = sheetproblem.cell_value(row,PROBLEMINDEX)
    problem_section = sheetproblem.cell_value(row,PROBLEMSECTION)
    problem_subsection = sheetproblem.cell_value(row,PROBLEMSUBSECTION)
    problem_unit = sheetproblem.cell_value(row,PROBLEMUNIT)
    problem_dir = sheetproblem.cell_value(row,PROBLEMDIR)
    problem_name = sheetproblem.cell_value(row,PROBLEMNAME)
    problem_sheet = sheetproblem.cell_value(row,PROBLEMSHEETNAME)
    problem_type = sheetproblem.cell_value(row,PROBLEMTYPE)
    problem_info.append({
        'idx':int(problem_idx),
        'section':problem_section,
        'subsection':problem_subsection,
        'unit':problem_unit,
        'dir':problem_dir,
        'filename':problem_name,
        'sheet':problem_sheet,
        'type':problem_type
    }) 
    return(problem_info[0])

def find_section_name(row_section,course_section):
    for course_sec_row in course_section:
        course_sec_row['section_name'] = course_sec_row['section_name'].rstrip()
        row_section['section'] = row_section['section'].rstrip()
        if course_sec_row['section_name']== row_section['section']:
            log.info ('found section: ' + (row_section['section'])+ ' in the exported course')
            selected_section = course_sec_row
            return selected_section

    error_msg = 'Pas de section: ' + (row_section['section']) + ' dans le cours exporté'
    return JsonResponse({"message" : error_msg})

def find_subsection_name(row_subsection,course_subsection,selected_section):
    for course_subsec_row in course_subsection:
        course_subsec_row['subsection_name'] = course_subsec_row['subsection_name'].rstrip()
        row_subsection['subsection'] = row_subsection['subsection'].rstrip()
        if course_subsec_row['subsection_name']== row_subsection['subsection']:
            if course_subsec_row['subsection_link'] in selected_section['assoc_subsection_url']:
                log.info ('found subsection: ' + (row_subsection['subsection'])+ ' in the exported course')
                selected_subsection = course_subsec_row
                return selected_subsection

    error_msg = 'Pas de sous-section: ' + (row_subsection['subsection']) + ' dans le cours exporté'
    return JsonResponse({"message" : error_msg})


def find_unit_name(row_unit,course_unit,selected_subsection,course_path):
    
    for course_unit_row in course_unit:
        course_unit_row['unit_name'] = course_unit_row['unit_name'].rstrip()
        row_unit['unit'] = row_unit['unit'].rstrip()
        if course_unit_row['unit_name'] == row_unit['unit']:
            if course_unit_row['unit_link'] in selected_subsection['assoc_unit_url']:
                log.info ('found unit: ' + row_unit['unit']+ ' in the exported course')
                tree = etree.parse(os.path.join(course_path,'vertical',course_unit_row['unit_link']+'.xml'))
                root = tree.getroot()
                new_problem_link = "problem" +  "{0:0=2d}".format(int(row_unit['idx']))
                etree.SubElement(root, 'problem',url_name=new_problem_link)
                doc = etree.ElementTree(root)
                doc.write(os.path.join(course_path,'vertical',course_unit_row['unit_link']+'.xml'), pretty_print=True, xml_declaration=False, encoding='utf-8')
                selected_unit = {'unit_link':course_unit_row['unit_link'],'unit_name':course_unit_row['unit_name'],'assoc_problem_url':new_problem_link}
                log.info('      added problem link: '+new_problem_link)

                return selected_unit

    error_msg = "Pas d'unité: " + (row_unit['unit']) + ' dans le cours exporté'
    return JsonResponse({"message" : error_msg})



def add_problem(problem_source_info,selected_unit,course_path):
    new_problem_link = "problem" +  "{0:0=2d}".format(int(problem_source_info['idx']))

    # if problem_source_info['type'] == 'multiple_choice':
    #     problem_instance = Problem_multichoice(problem_source_info)
    # elif problem_source_info['type'] == 'droplist':
    #     problem_instance = Problem_droplist(problem_source_info)
    # elif problem_source_info['type'] == 'checkbox':

    if problem_source_info['type'] == 'checkbox':
        problem_instance = Problem_checkbox(problem_source_info)
    # elif problem_source_info['type'] == 'fill_blank':
    #     problem_instance = Problem_fillblank(problem_source_info)
    else:
        error_msg = 'Argument "type de problème" est manquant'
        return JsonResponse({"message" : error_msg})


    problem_instance.create_file(new_problem_link+'.xml',course_path)


def search_problem_in_course(row_from_excel,course_structure,course_path):
    
    selected_section = find_section_name(row_from_excel,course_structure.sections())
    selected_subsection = find_subsection_name(row_from_excel,course_structure.subsections(),selected_section)
    selected_unit = find_unit_name(row_from_excel,course_structure.units(),selected_subsection,course_path)
    add_problem(row_from_excel,selected_unit,course_path)


@require_http_methods(["POST"])
@login_required
def convert_to_tarfile_bvt(request):

    course_title = request.POST['courseTitle'] 
    degree_of_cert = request.POST['degreeOfCert'] 
    wbinput = load_workbook(request.FILES['excelFile'], data_only=True)
    sheet = wbinput.worksheets[0]

    # CREATE A DICT FROM DATA
    data = dict()
    row_count = sheet.max_row-1

    # CREATE DICT FOR list_correct_answer JSON
    data_list_correct_answer = dict()
    for i in range(row_count):
        try:
            data_current=dict()
            if sheet.cell(row=i+2, column=5).value != None:
                data_current["question"] = sheet.cell(row=i+2, column=5).value

                # answer list
                list_answer = []
                list_answer.append(sheet.cell(row=i+2, column=6).value)
                list_answer.append(sheet.cell(row=i+2, column=8).value)
                list_answer.append(sheet.cell(row=i+2, column=10).value)
                list_answer.append(sheet.cell(row=i+2, column=12).value)
                data_current["answer_list"]=list_answer

                # Question type :- We only want checkbox, multiple_choice will gave a hint to the student
                data_current["type"] = 'checkbox'

                # Answer list
                list_correct_answer = []
                if sheet.cell(row=i+2, column=7).value == 1:
                    list_correct_answer.append('0')
                if sheet.cell(row=i+2, column=9).value == 1:
                    list_correct_answer.append('1')
                if sheet.cell(row=i+2, column=11).value == 1:
                    list_correct_answer.append('2')
                if sheet.cell(row=i+2, column=13).value == 1:
                    list_correct_answer.append('3')
                data_current["correct_answer_index"] = list_correct_answer

                problem_index = sheet.cell(row=i+2, column=16).value.split(' ')[1]

                if len(problem_index) == 1 :
                    problem_index = '0' + problem_index

                data_list_correct_answer['problem'+problem_index] = list_correct_answer

                # max_attempts
                # data_current['max_attempts'] = 1

                # hint 
                if sheet.cell(row=i+2, column=14).value != None:
                    data_current["hint"] = sheet.cell(row=i+2, column=14).value
                else:
                    data_current["hint"] = None

                # section, subsection and, unit
                data_current["title"]=course_title
                data_current["section"]=sheet.cell(row=i+2, column=17).value
                data_current["subsection"]=sheet.cell(row=i+2, column=18).value
                data_current["unit"]= sheet.cell(row=i+2, column=19).value

                if sheet.cell(row=i+2, column=15).value != None:

                    image_name= sheet.cell(row=i+2, column=15).value
                    log.info('image_name')
                    log.info(image_name)
                    try:
                        image_name = image_name.replace("<<< Fig/  ", "")
                    except: 
                        log.info("error image 1")
                        index = i+2
                        error_msg = "Problem detected regarding image reference at line " + str(index)
                        return JsonResponse({"message" : error_msg})

                    try:
                        image_name = image_name.replace(" >>>", ".jpg")
                    except: 
                        log.info("error image 2")
                        index = i+2
                        error_msg = "Problem detected regarding image reference at line " + str(index)
                        return JsonResponse({"message" : error_msg})

                    data_current["image"]= "https://bvt.koa.qualif.dev/media/microsites/bvt/img_bvt_converter/" + image_name
                
                else:
                    data_current["image"]= None


                # displ_name
                data_current["displ_name"] = sheet.cell(row=i+2, column=16).value

                # save each
                data[i+1]= data_current
        except:
            index = i+2
            error_msg = "Une erreur a été détectée à la ligne " + str(index) + ", vérifiez le fichier importé ou contactez le service technique de WeUp Learning"
            return JsonResponse({"message" : error_msg})

    # WRITE FILE " questions.xls "
    wb_question = xlwt.Workbook()


    # WRITE JSON list_correct_answer
    json_file_name = 'list_corrected_answer_' + str(course_title).replace(' ', '_') +'.json'
    log.info(json_file_name)
    with open('/edx/var/edxapp/media/microsites/bvt/answers_lists_files/'+ json_file_name, 'w') as json_file :
        json.dump(data_list_correct_answer ,json_file)
        log.info('------------> Save corrected answers')


    for i, question in enumerate(data):

        headers = ["problem_display_name", "grade_weight", "max_attempts", "hint", "subquestion", "choice", "answer", "image", "displ_name"]
        # new sheet
        ws = wb_question.add_sheet("Q "+str(question))

        # headers
        for j, header in enumerate(headers):
            ws.write(0, j, header)

        # question
        ws.write(1, 4, data[question]['question'])

        # unit-name
        ws.write(1, 0, data[question]['displ_name'])
        # ws.write(1, 0, data[question]['unit'])

        # max_attempts
        # if data[question]['max_attempts'] != None:
        #     ws.write(1, 2, data[question]['max_attempts'])

        # hint
        if data[question]['hint'] != None:
            ws.write(1, 3, data[question]['hint'])

        # answer list and correct answer
        for k, answer in enumerate(data[question]["answer_list"]):
            ws.write(k+1, 5, answer)

            if str(k) in data[question]['correct_answer_index'] :
                ws.write(k+1, 6, "t")
            else: 
                ws.write(k+1, 6, "f")

        # image
        log.info(data[question]["image"])
        if data[question]["image"] != None:
            ws.write(1, 7, data[question]["image"])

        # Display_name
        ws.write(1, 8, data[question]["unit"])

    filename_output= "questions.xls"
    filepath1 = '/edx/var/edxapp/media/microsites/bvt/edx_converter/problem_source/{}'.format( filename_output)

    wb_question.save(filepath1)

    # Write " course_info.xlsx "
    wb_course_info = xlwt.Workbook()
    component_types = ["text", "video", "problem" ]

    ws = wb_course_info.add_sheet("coursestructure")
    ws.write(0, 0, "idx")
    ws.write(0, 1, "section")
    ws.write(0, 2, "subsection")
    ws.write(0, 3, "unit")
    ws.write(0, 4, "component")
    ws.write(0, 5, "component_types")
    for i, question in enumerate(data):
        ws.write(i+1, 0, question)
        ws.write(i+1, 1, data[question]['section'])
        ws.write(i+1, 2, data[question]['subsection'])
        ws.write(i+1, 3, data[question]['unit'])
        ws.write(i+1, 4, question)
        ws.write(i+1, 5, component_types[2])

    ws = wb_course_info.add_sheet(component_types[2])
    ws.write(0, 0, "index")
    ws.write(0, 1, "sectionname")
    ws.write(0, 2, "subsectionname")
    ws.write(0, 3, "unitname")
    ws.write(0, 4, "file_dir")
    ws.write(0, 5, "filename")
    ws.write(0, 6, "sheet_name")
    ws.write(0, 7, "problem_name")
    ws.write(0, 8, "problem_type")

    for i, question in enumerate(data):
        ws.write(i+1, 0, question)
        ws.write(i+1, 1, data[question]['section'])
        ws.write(i+1, 2, data[question]['subsection'])
        ws.write(i+1, 3, data[question]['unit'])
        # directory name
        ws.write(i+1, 4, "{}_source".format(component_types[2]))
        # file name
        ws.write(i+1, 5, "questions.xls")
        # sheet name
        ws.write(i+1, 6, "Q "+str(question))
        # probleme name
        ws.write(i+1, 7, question)
        ws.write(i+1, 8, data[question]['type'])

    ws = wb_course_info.add_sheet(component_types[1])
    ws.write(0, 0, "idx")
    ws.write(0, 1, "section")
    ws.write(0, 2, "subsection")
    ws.write(0, 3, "unit")
    ws.write(0, 4, "video_url")
    ws.write(0, 5, "video_name")
    ws.write(0, 6, "file_dir")
    ws.write(0, 7, "en_sub")
    ws.write(0, 8, "jp_sub")

    ws = wb_course_info.add_sheet(component_types[0])
    ws.write(0, 0, "idx")
    ws.write(0, 1, "section")
    ws.write(0, 2, "subsection")
    ws.write(0, 3, "unit")
    ws.write(0, 4, "text_name")
    ws.write(0, 5, "file_dir")
    ws.write(0, 6, "file_name")

    ws = wb_course_info.add_sheet("upload_list")
    ws.write(0, 0, "id")
    ws.write(0, 1, "file_dir")
    ws.write(0, 2, "filename")
    ws.write(0, 3, "title")
    ws.write(0, 4, "description")
    ws.write(0, 5, "keyword")
    ws.write(0, 6, "privacy_status")

    ws = wb_course_info.add_sheet("caption_list")
    ws.write(0, 0, "id")
    ws.write(0, 1, "file_dir")
    ws.write(0, 2, "filename")
    ws.write(0, 3, "lang")
    ws.write(0, 4, "name")
    ws.write(0, 5, "video_id")
    ws.write(0, 6, "remark")

    ws = wb_course_info.add_sheet("thumbnail_list")
    ws.write(0, 0, "id")
    ws.write(0, 1, "file_dir")
    ws.write(0, 2, "filename")
    ws.write(0, 3, "video_id")
    ws.write(0, 4, "remark")

    filename_output2= "course_info.xls"
    filepath2 = '/edx/var/edxapp/media/microsites/bvt/edx_converter/{}'.format(filename_output2)
    wb_course_info.save(filepath2)

    STRUCID = 0
    STRUCSECTION = 1
    STRUCSUBSECTION = 2
    STRUCUNIT = 3
    STRUCCOMPONENT = 4
    STRUCTYPECOMPONENT = 5

    # KOA path
    course_path= '/edx/var/edxapp/media/microsites/bvt/edx_converter/course'
    xlsmPath = "/edx/var/edxapp/media/microsites/bvt/edx_converter/course_info.xls"
    wb = xlrd.open_workbook(xlsmPath)

    sheetstruc = wb.sheet_by_name("coursestructure")
    # sheetvideo = wb.sheet_by_name(VIDEOSHEET)
    sheetproblem = wb.sheet_by_name("problem")
    # sheethtml = wb.sheet_by_name(HTMLSHEET)


    class Course_extraction:

        def __init__(self):
            if not os.path.exists(os.path.join(course_path,"chapter")):
                os.makedirs(os.path.join(course_path,"chapter"))
            if not os.path.exists(os.path.join(course_path,"sequential")):
                os.makedirs(os.path.join(course_path,"sequential"))
            if not os.path.exists(os.path.join(course_path,"vertical")):
                os.makedirs(os.path.join(course_path,"vertical"))
            if not os.path.exists(os.path.join(course_path,"video")):
                os.makedirs(os.path.join(course_path,"video"))
            if not os.path.exists(os.path.join(course_path,"problem")):
                os.makedirs(os.path.join(course_path,"problem"))
            if not os.path.exists(os.path.join(course_path,"html")):
                os.makedirs(os.path.join(course_path,"html"))
            if not os.path.exists(os.path.join(course_path,"static")):
                os.makedirs(os.path.join(course_path,"static"))

            self.section_path = os.path.join(course_path,'chapter')
            self.subsection_path = os.path.join(course_path,'sequential')
            self.unit_path = os.path.join(course_path,'vertical')
            self.problme_path = os.path.join(course_path,'problem')
            self.course = os.path.join(course_path,'course')


        def course_(self):
            section_file = os.path.join(self.course,'course.xml')
            tree = etree.parse(section_file)
            root = tree.getroot()
            section_ls = root.findall('.chapter')
            temp_url = []
            for section in section_ls:
                temp_url.append(section.get('url_name'))
            self.section_url = {'section_url':temp_url}

            return(self.section_url)


        def sections(self):
            sections_files = os.listdir(self.section_path)
            self.all_section = []
            for section_file in sections_files:
                tree = etree.parse(os.path.join(self.section_path,section_file))
                root = tree.getroot()
                section_name = root.get('display_name')
                section_link = section_file.replace('.xml', '')
                subsection_objs = root.findall(".sequential")
                subsection_url = []
                for subsection_obj in subsection_objs:    
                    subsection_url.append(subsection_obj.get('url_name'))

                self.all_section.append({'section_link':section_link,
                    'section_name':section_name,
                    'assoc_subsection_url':subsection_url})
            return(self.all_section)


        def subsections(self):
            subsections_files = os.listdir(self.subsection_path)
            self.all_subsection = []
            for subsection_file in subsections_files:
                tree = etree.parse(os.path.join(self.subsection_path,subsection_file))
                root = tree.getroot()
                subsection_name = root.get('display_name')
                subsection_link = subsection_file.replace( '.xml', '')
                unit_objs = root.findall(".vertical")
                unit_url = []
                for unit_obj in unit_objs:
                    unit_url.append(unit_obj.get('url_name'))

                self.all_subsection.append({'subsection_link':subsection_link,
                    'subsection_name':subsection_name,
                    'assoc_unit_url':unit_url})
            return(self.all_subsection)


        def units(self):
            units_files = os.listdir(self.unit_path)
            self.all_unit = []
            for unit_file in units_files:
                tree = etree.parse(os.path.join(self.unit_path,unit_file))
                root = tree.getroot()
                unit_name = root.get('display_name')
                unit_link = unit_file.replace('.xml', '')
                self.all_unit.append({'unit_link':unit_link,'unit_name':unit_name})
            return(self.all_unit)


    def create_course(course_title, degree_of_cert):
        """
        create course.xml file
        """
        from_course = Course_extraction()
        list_section = from_course.course_()
        tree = etree.parse(os.path.join(course_path,'course','course.xml'))
        root = tree.getroot()

        # MODIFY DISPLAY_NAME
        root.set('display_name', course_title)

        if str(degree_of_cert) == "false" :
            # root.set('degree_of_certainty', '')
            root.set('cert_html_view_enabled', 'false')
        else:
            root.set('cert_html_view_enabled', 'true')

        currentsection = ''
        section_idx = 1

        for row in range(1, sheetstruc.nrows):

            if currentsection != sheetstruc.cell_value(row, STRUCSECTION ):
                currentsection = sheetstruc.cell_value(row, STRUCSECTION )
                urlName = "section" +  "{0:0=2d}".format(section_idx)
                if urlName not in list_section['section_url']:
                    log.info('no section: "'+ urlName +'"" in course. Add link to course.xml')
                    etree.SubElement(root, 'chapter',url_name=urlName)
                else:
                    log.info('section: "'+ urlName +'" exists in course.')
                section_idx+=1

        # Add completion section here
        urlName = "d4a9392d58cb49fba5afe41b33aa8f9e"
        if urlName not in list_section['section_url']:
            log.info('no section: "'+ urlName +'" in course. Add link to course.xml')
            etree.SubElement(root, 'chapter',url_name=urlName)

        doc = etree.ElementTree(root)
        doc.write(os.path.join(course_path,'course','course.xml'), pretty_print=True, xml_declaration=False, encoding='utf-8')


    def create_section():
        """
        Creates a section file
        """
        currentsection = sheetstruc.cell_value(1, STRUCSECTION )
        currentsubsection = sheetstruc.cell_value(1,STRUCSUBSECTION)
        section_idx = 1
        subsection_idx = 1
        filename = 'section' +  '{0:0=2d}'.format(section_idx) + '.xml'
        page = etree.Element('chapter', display_name= currentsection)
        subsection_url_name = 'subsection' +  '{0:0=2d}'.format(subsection_idx) 
        etree.SubElement(page, 'sequential',url_name=subsection_url_name)
        subsection_idx += 1
        log.info('added new section: "'+ filename +'" file at chapter directory')
        log.info('      added new subsection link"'+ subsection_url_name +'"" in section:' +filename )

        for row in range(2, sheetstruc.nrows):

            if currentsection != sheetstruc.cell_value(row, STRUCSECTION ):
                doc = etree.ElementTree(page)
                doc.write(os.path.join(course_path,'chapter',filename), pretty_print=True, xml_declaration=False, encoding='utf-8')
                log.info('added new section: "'+ filename +'" file at chapter directory')
                section_idx +=1
                currentsection = sheetstruc.cell_value(row, STRUCSECTION )
                currentsubsection = sheetstruc.cell_value(row,STRUCSUBSECTION)
                filename = 'section' +  '{0:0=2d}'.format(section_idx) + '.xml'
                page = etree.Element('chapter', display_name= currentsection)
                subsection_url_name = 'subsection' +  '{0:0=2d}'.format(subsection_idx) 
                etree.SubElement(page, 'sequential',url_name=subsection_url_name)
                subsection_idx += 1
                log.info('      added new subsection link"'+ subsection_url_name +'"" in file: ' +filename )
            else:
                if currentsubsection != sheetstruc.cell_value(row,STRUCSUBSECTION):
                    currentsubsection = sheetstruc.cell_value(row,STRUCSUBSECTION)
                    subsection_url_name = 'subsection' +  '{0:0=2d}'.format(subsection_idx) 
                    etree.SubElement(page, 'sequential',url_name=subsection_url_name)
                    subsection_idx += 1
                    log.info('      added new subsection link "'+ subsection_url_name +'"" in file: ' +filename )

        doc = etree.ElementTree(page)
        doc.write(os.path.join(course_path,'chapter',filename), pretty_print=True, xml_declaration=False, encoding='utf-8')
        log.info('added new section: "'+ filename +'" file at chapter directory')

        # Add a xml in chapter for completion block
        page_completion = etree.Element('chapter', display_name="Completion")
        subsection_url_name = '310e3b2e46784b028a88a55034f35ce7' 

        etree.SubElement(page_completion, 'sequential',url_name=subsection_url_name)

        doc = etree.ElementTree(page_completion)
        doc.write(os.path.join(course_path,'chapter','d4a9392d58cb49fba5afe41b33aa8f9e.xml'), pretty_print=True, xml_declaration=False, encoding='utf-8')
        log.info('added new section: "'+ filename +'" file at chapter directory')


    def create_subsection():
        """
        Creates subsection files
        """
        currentsubsection = sheetstruc.cell_value(1,STRUCSUBSECTION)
        currentunit = sheetstruc.cell_value(1, STRUCUNIT)
        subsection_idx = 1
        unit_idx = 1
        filename = 'subsection' +  '{0:0=2d}'.format(subsection_idx) + '.xml'
        page = etree.Element('sequential', display_name= currentsubsection, format="Exam", graded="true")
        unit_url_name = 'unit' +  '{0:0=2d}'.format(subsection_idx) 
        etree.SubElement(page, 'vertical',url_name=unit_url_name)
        unit_idx += 1
        log.info('added new subsection: "'+ filename +'" file at sequential directory')
        log.info('      added new unit link "'+ unit_url_name +'"" in subsection:' +filename )

        for row in range(2, sheetstruc.nrows):

            if currentsubsection != sheetstruc.cell_value(row, STRUCSUBSECTION ):
                doc = etree.ElementTree(page)
                doc.write(os.path.join(course_path,'sequential',filename), pretty_print=True, xml_declaration=False, encoding='utf-8')
                log.info('added new subsection: "'+ filename +'" file at sequential directory')
                subsection_idx += 1
                currentsubsection = sheetstruc.cell_value(row, STRUCSUBSECTION )
                currentunit = sheetstruc.cell_value(row,STRUCUNIT)
                filename = 'subsection' +  '{0:0=2d}'.format(subsection_idx) + '.xml'
                page = etree.Element('sequential', display_name= currentsubsection, format="Exam", graded="true")
                unit_url_name = 'unit' +  '{0:0=2d}'.format(unit_idx) 
                etree.SubElement(page, 'vertical',url_name=unit_url_name)
                unit_idx += 1
                log.info('      added new unit "'+  unit_url_name +'" in file: ' +filename )

            else:
                if currentunit != sheetstruc.cell_value(row,STRUCUNIT):
                    currentunit = sheetstruc.cell_value(row,STRUCUNIT)
                    unit_url_name = 'unit' +  '{0:0=2d}'.format(unit_idx) 
                    etree.SubElement(page, 'vertical',url_name=unit_url_name)
                    unit_idx += 1
                    log.info('      added new unit "'+ unit_url_name +'" in file: ' +filename )

        doc = etree.ElementTree(page)
        doc.write(os.path.join(course_path,'sequential',filename), pretty_print=True, xml_declaration=False, encoding='utf-8')
        log.info('added new subsection: "'+ filename +'" file at sequential directory')

        # Add subsection for completion block
        subsection_completion = etree.Element('sequential', display_name="Completion")
        unit_url_name_compl = "9a65db5a76504695ba4aab3889de22b9"

        etree.SubElement(subsection_completion, 'vertical', url_name=unit_url_name_compl)

        doc = etree.ElementTree(subsection_completion)
        doc.write(os.path.join(course_path,'sequential','310e3b2e46784b028a88a55034f35ce7.xml'), pretty_print=True, xml_declaration=False, encoding='utf-8')


    def create_unit():
        """
        Creates unit files
        """
        currentunit = sheetstruc.cell_value(1, STRUCUNIT)
        unit_idx = 1
        filename = 'unit' +  '{0:0=2d}'.format(unit_idx) +'.xml'
        page = etree.Element('vertical', display_name= currentunit)
        log.info('added new unit: "'+ filename +'" file at vertical directory')

        for row in range(2, sheetstruc.nrows):
            if currentunit != sheetstruc.cell_value(row, STRUCUNIT):
                doc = etree.ElementTree(page)
                doc.write(os.path.join(course_path,'vertical',filename), pretty_print=True, xml_declaration=False, encoding='utf-8')
                log.info('added new unit: "'+ filename +'" file at vertical directory')
                unit_idx +=1
                currentunit = sheetstruc.cell_value(row, STRUCUNIT )
                filename = 'unit' +  '{0:0=2d}'.format(unit_idx) + '.xml'
                page = etree.Element('chapter', display_name= currentunit)

        doc = etree.ElementTree(page)
        doc.write(os.path.join(course_path,'vertical',filename), pretty_print=True, xml_declaration=False, encoding='utf-8')
        log.info('added new unit: "'+ filename +'" file at vertical directory')

        # Add unit for completion block
        completion_unit = 'completion_unit'
        page = etree.Element('vertical', display_name= completion_unit)

        html_id = '688887f23f184366934880290c4e9731'
        etree.SubElement(page, 'html', url_name=html_id)

        doc = etree.ElementTree(page)
        doc.write(os.path.join(course_path,'vertical','9a65db5a76504695ba4aab3889de22b9.xml'), pretty_print=True, xml_declaration=False, encoding='utf-8')


    def add_component():
        '''
        Start adding component with respect to macro excel
        '''
        problem_idx = 1
        for row in range(1, sheetstruc.nrows):
            comp_type = sheetstruc.cell_value(row,STRUCTYPECOMPONENT)
            if comp_type == 'problem':
                search_problem_in_course(problem_excel2list(problem_idx,sheetproblem),Course_extraction(),course_path)
                problem_idx +=1


    def make_tarfile():
        '''
        Compile the targeted directories into a tar.gz file (open edX readable)
        '''
        tar_path = '/edx/var/edxapp/media/microsites/bvt/course'
        # compress course content in a targz file and ready to import.
        log.info("file is being compressed as tar.gz ")
        with tarfile.open(tar_path + '.tar.gz', 'w:gz') as tar:
            for f in os.listdir(course_path):
                tar.add(course_path + "/" + f, arcname=os.path.basename(f))
            tar.close()
        log.info('uploadable file is created at ' + tar_path + '.tar.gz')


    def update_json():
        '''
        Write into policy.json to update advanced parameters
        '''
        json_path = course_path + "/policies/course/policy.json"
        # UPDATE DATE HERE
        today = date.today()
        formatedDate = today.strftime('%Y-%m-%d')

        with open(json_path, "r") as json_file:
            update_title = json.load(json_file)
            update_title["course/course"]["display_name"]= course_title
            update_title["course/course"]["course_image"]= "vignette_test.png"
            update_title["course/course"]["start"]= str(formatedDate) +"T00:00:00Z"

            # if str(degree_of_cert) == "false":
            #     update_title["course/course"]["degree_of_certainty"]["_on"]= False
            # else :
            #     update_title["course/course"]["degree_of_certainty"]["_on"]= True

        os.remove(json_path)
        with open(json_path, "w") as f:
            json.dump(update_title, f ,indent=4)


    def delete_directories():
        '''
        Clean the implemented directories to start over the main process 
        '''
        dir_list = ["chapter", "problem", "sequential", "vertical", "video"]
        for directory in dir_list:
            dir_path = "/edx/var/edxapp/media/microsites/bvt/edx_converter/course/" + directory
            try:
                shutil.rmtree(dir_path)
            except OSError as e:
                print("Error: %s : %s" % (dir_path, e.strerror))

    log.info ('Create course outline')
    create_course(course_title, degree_of_cert)
    log.info('Apply new title and degree of certainty')
    update_json()
    create_section()
    create_subsection()
    create_unit()
    log.info ('Add course contents')
    add_component()
    make_tarfile()
    log.info("Delete files after tarfile is generated")
    delete_directories()

    response = HttpResponse()
    return response
